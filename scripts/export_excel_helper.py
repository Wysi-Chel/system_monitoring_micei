import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(fill_type="solid", fgColor="C6E0B4")
HEADER_FONT = Font(name="Verdana", bold=True, color="000000", size=12)
BODY_FONT = Font(name="Verdana", color="000000", size=12)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(vertical="center")
BODY_CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
BODY_RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center")
THIN_GRID = Side(style="thin", color="D9D9D9")
HEADER_BOTTOM = Side(style="medium", color="000000")
LIGHT_BORDER = Border(left=THIN_GRID, right=THIN_GRID, top=THIN_GRID, bottom=THIN_GRID)
HEADER_BORDER = Border(left=THIN_GRID, right=THIN_GRID, top=THIN_GRID, bottom=HEADER_BOTTOM)

DATE_HEADERS = {"Date", "Transaction Date", "Date Created", "Date Resolved"}
CENTER_HEADERS = {"Date", "Transaction Date", "Date Created", "Date Resolved", "Branch", "Dealers", "Department", "Module", "Amount", "Status", "Alert", "Action"}


def sanitize_sheet_name(sheet_name: str) -> str:
    invalid_chars = "\\/*?:[]"
    cleaned = "".join(" " if char in invalid_chars else char for char in sheet_name).strip()
    return (cleaned or "Sheet1")[:31]


def coerce_amount(value):
    if value in ("", None):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    try:
        return float(str(value))
    except ValueError:
        return value


def calculate_width(values) -> float:
    max_length = max((len(str(value)) for value in values if value not in (None, "")), default=10)
    return min(max(max_length + 5, 13), 42)


def should_preserve_token(token: str) -> bool:
    core = re.sub(r"^[^\w]+|[^\w#/\-()]+$", "", token, flags=re.UNICODE)
    if not core:
        return False

    if any(char.isdigit() for char in core):
        return True

    if any(char in "#/-()" for char in core):
        return True

    letters_only = re.sub(r"[^A-Za-z]", "", core)
    return bool(letters_only) and 2 <= len(letters_only) <= 6 and letters_only.isupper()


def lowercase_token(token: str) -> str:
    return token.lower()


def capitalize_first_letter(text: str) -> str:
    for index, char in enumerate(text):
        if char.isalpha():
            return text[:index] + char.upper() + text[index + 1 :]
    return text


def sentence_case_text(value):
    if not isinstance(value, str):
        return value

    value = re.sub(r"\s+", " ", value.strip())
    if not value:
        return ""

    parts = re.split(r"([.!?]+\s*)", value)
    normalized_parts = []

    for index, part in enumerate(parts):
        if index % 2 == 1:
            normalized_parts.append(part)
            continue

        tokens = re.split(r"(\s+)", part)
        normalized_tokens = []

        for token in tokens:
            if not token or token.isspace():
                normalized_tokens.append(token)
            elif should_preserve_token(token):
                normalized_tokens.append(token)
            else:
                normalized_tokens.append(lowercase_token(token))

        normalized_parts.append(capitalize_first_letter("".join(normalized_tokens)))

    return "".join(normalized_parts)


def format_short_date(value: datetime) -> str:
    return f"{value.month}/{value.day}/{value.year}"


def format_timestamp_display(value: datetime) -> str:
    hour = value.hour % 12 or 12
    suffix = "AM" if value.hour < 12 else "PM"
    return f"{value.month}/{value.day}/{value.year} {hour}:{value.minute:02d} {suffix}"


def display_value(value, header):
    if value in ("", None):
        return ""

    if header in DATE_HEADERS and isinstance(value, datetime):
        return format_short_date(value)

    if header == "Encoded At" and isinstance(value, datetime):
        return format_timestamp_display(value)

    if header == "Amount" and isinstance(value, (int, float)):
        return f"{value:,.2f}"

    return str(value)


def parse_short_date(value):
    if value in ("", None):
        return None

    if isinstance(value, datetime):
        return value

    try:
        return datetime.strptime(str(value), "%m/%d/%Y")
    except ValueError:
        try:
            return datetime.strptime(str(value), "%-m/%-d/%Y")
        except ValueError:
            return value


def parse_timestamp(value):
    if value in ("", None):
        return None

    if isinstance(value, datetime):
        return value

    try:
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return value


def format_header_row(worksheet) -> None:
    worksheet.row_dimensions[1].height = 34

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER


def format_body_cell(cell, header) -> None:
    cell.font = BODY_FONT
    cell.border = LIGHT_BORDER

    if header == "Amount":
        cell.alignment = BODY_RIGHT_ALIGNMENT
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"
        return

    if header in DATE_HEADERS and isinstance(cell.value, datetime):
        cell.alignment = BODY_CENTER_ALIGNMENT
        cell.number_format = "m/d/yyyy"
        return

    if header == "Encoded At" and isinstance(cell.value, datetime):
        cell.alignment = BODY_CENTER_ALIGNMENT
        cell.number_format = "m/d/yyyy h:mm AM/PM"
        return

    if header in CENTER_HEADERS:
        cell.alignment = BODY_CENTER_ALIGNMENT
        return

    cell.alignment = BODY_ALIGNMENT


def apply_column_widths(worksheet, headers) -> None:
    for column_index, header in enumerate(headers, start=1):
        values = [
            display_value(worksheet.cell(row=row_index, column=column_index).value, header)
            for row_index in range(1, worksheet.max_row + 1)
        ]
        worksheet.column_dimensions[get_column_letter(column_index)].width = calculate_width(values)


def main() -> int:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: export_excel_helper.py <payload.json> <output.xlsx>")

    payload_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    headers = payload.get("headers", [])
    rows = payload.get("rows", [])

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sanitize_sheet_name(payload.get("sheet_name", "Monitoring Summary"))
    worksheet.sheet_properties.tabColor = "70AD47"
    worksheet.sheet_view.zoomScale = 90
    worksheet.sheet_view.zoomScaleNormal = 90

    worksheet.append(headers)
    format_header_row(worksheet)

    for raw_row in rows:
        row = [sentence_case_text(value) for value in raw_row]

        for index, header in enumerate(headers):
            if index >= len(row):
                continue

            if header in DATE_HEADERS:
                row[index] = parse_short_date(row[index])
                continue

            if header == "Amount":
                row[index] = coerce_amount(row[index])
                continue

            if header == "Encoded At":
                row[index] = parse_timestamp(row[index])

        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.selection[0].activeCell = "A2"
    worksheet.sheet_view.selection[0].sqref = "A2"

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 22
        for column_index, header in enumerate(headers, start=1):
            format_body_cell(worksheet.cell(row=row_index, column=column_index), header)

    apply_column_widths(worksheet, headers)

    workbook.save(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
